from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q, Sum
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from rest_framework.pagination import PageNumberPagination
from rest_framework import generics, status, viewsets, mixins, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, action
from django.templatetags.static import static
from decimal import Decimal
from django.db.models.signals import pre_delete, post_delete
from django.dispatch import receiver
import datetime
import random
from django.utils import timezone
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .decorators import moderator_role_required
from django.urls import reverse
from django.contrib import messages
from .firebase_messaging import send_push_notification
from django.core.mail import send_mail
from django.conf import settings
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from django.utils.decorators import method_decorator
import logging
from .serializers import (
    UserSerializer,
    CategorySerializer,
    ServiceSerializer,
    DoctorSerializer,
    DoctorCreateUpdateSerializer,
    DoctorScheduleSerializer,
    DoctorPermissionSerializer,
    LaserUsageSerializer,
    TreatmentSerializer,
    DiscountCodeSerializer,
    DiscountBannerSerializer,
    MainBannerSerializer,
    ReservationSerializer,
    DoctorServiceSerializer,
    ExcelReservationSerializer,
    ModeratorSerializer,
    NotificationSerializer,
    UserRegisterSerializer,
    UserLoginSerializer,
    ActivateUserSerializer,
    UserCashbackSerializer,
    UserCashbackDetailSerializer,
    CashbackHistorySerializer,
    ForgotPasswordSerializer,
    ReservationTreatmentSerializer,
    VerifyOtpSerializer,
    UpdatePasswordSerializer,
    ResetPasswordSerializer,
    ReservationDetailSerializer,
    ChangePasswordSerializer,
)
from .models import (User, Category, Service, Doctor, DoctorSchedule, DoctorPermission,
                     LaserUsage, Treatment, DiscountCode, DiscountBanner,
                     MainBanner, Reservation, UserCashback, CashbackHistory, Payment,
                     DoctorService, Moderator, Notification, HistoryLog)


def get_first_true_role_page(moderator):
    role_to_url = [
        ("can_view_statistics", "statistics_view"),
        ("can_view_calendar", "reservations_view"),
        ("can_view_users", "users_view"),
        ("can_view_services", "services_view"),
        ("can_view_notification", "notification_view"),
        ("can_view_treatment", "treatment_view"),
        ("can_view_cashback", "cashback_view"),
        ("can_view_pages", "pages_view"),
        ("can_view_discount_code", "discount_code_view"),
        ("can_view_doctors", "doctors_view"),
        ("can_view_history", "history_view"),
        ("can_view_moderator", "moderator_view"),
        ("can_view_laser", "laser_view"),
        ("can_view_payment_acceptance", "payment_acceptance_view"),
        ("can_view_doctor_payment", "doctor_payment_view"),
        ("can_view_excel", "excel_view"),
    ]
    for role_field, url_name in role_to_url:
        if getattr(moderator, role_field, False):
            return reverse(url_name)
    return None


def login_submit(request):
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        remember_me = request.POST.get("remember_me", "0")

        try:
            moderator = Moderator.objects.get(email=email, password=password)
        except Moderator.DoesNotExist:
            messages.error(request, "E-poçt və ya parol səhvdir.")
            return redirect("login_view")

        if moderator.status != "active":
            messages.error(
                request, "Moderator deaktivdir. Zəhmət olmasa sistem administratoru ilə əlaqə saxlayın.")
            return redirect("login_view")

        request.session["moderator_id"] = moderator.id
        if remember_me == "1":
            request.session.set_expiry(60 * 60 * 24 * 30)
        else:
            request.session.set_expiry(0)

        first_page = get_first_true_role_page(moderator)
        if first_page:
            return redirect(first_page)
        else:
            messages.error(
                request, "Moderatorun heç bir səhifəyə icazəsi yoxdur.")
            return redirect("login_view")
    return redirect("login_view")


def logout_view(request):
    if "moderator_id" in request.session:
        del request.session["moderator_id"]
    return redirect("login_view")


def login_view(request):
    moderator_id = request.session.get("moderator_id")
    if moderator_id:
        try:
            moderator = Moderator.objects.get(id=moderator_id)
            if moderator.status == "active":
                first_page = get_first_true_role_page(moderator)
                if first_page:
                    return redirect(first_page)
                else:
                    messages.error(
                        request, "Moderatorun heç bir səhifəyə icazəsi yoxdur.")
                    return redirect("login_view")
            else:
                messages.error(
                    request, "Moderator deaktivdir. Yenidən daxil olmağa çalışın.")
                return render(request, "index.html")
        except Moderator.DoesNotExist:
            pass

    return render(request, "index.html")


@moderator_role_required('can_view_users')
def users_view(request):
    search_query = request.GET.get('search', '')
    user_list = User.objects.all().order_by('-created_at')

    if search_query:
        user_list = user_list.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    page = request.GET.get('page', 1)
    paginator = Paginator(user_list, 10)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    context = {
        'users': users,
        'search_query': search_query,
    }
    return render(request, "users.html", context)


class UserListCreateAPIView(generics.ListCreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)

        moderator_id = request.session.get('moderator_id', None)
        if moderator_id:
            try:
                moderator = Moderator.objects.get(id=moderator_id)
                HistoryLog.objects.create(
                    moderator=moderator,
                    action_type='user_created'
                )
            except Moderator.DoesNotExist:
                pass

        return response


class UserRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

logger = logging.getLogger(__name__)
class UserRegisterAPIView(generics.CreateAPIView):
    serializer_class = UserRegisterSerializer
    queryset = User.objects.all()

    def perform_create(self, serializer):
        user = serializer.save()
        subject = "Your OTP Code"
        message = f"Your OTP code is: {user.otp_code}"
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [user.email]
        try:
            send_mail(subject, message, from_email, recipient_list)
        except Exception as e:
            logger.error("Error sending OTP email: %s", e)
        return user


class UserLoginAPIView(APIView):
    serializer_class = UserLoginSerializer

    @swagger_auto_schema(
        operation_summary="User login",
        operation_description=(
            "Bu endpoint istifadəçinin email və şifrəsi ilə daxil olmasına imkan verir."
        ),
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email", "password"],
            properties={
                "email": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    format=openapi.FORMAT_EMAIL,
                    description="User-in email ünvanı",
                    example="example@gmail.com"
                ),
                "password": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="User-in parolu",
                    example="123456"
                ),
                "fcm_token": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Firebase FCM token (optional)",
                    example="fcm_token_example_ABC123"
                ),
            }
        ),
        responses={
            200: openapi.Response(
                description="Uğurlu login",
                examples={
                    "application/json": {
                        "detail": "Login successful",
                        "user_id": 10
                    }
                }
            ),
            400: openapi.Response(
                description="Yanlış məlumat və ya istifadəçi aktiv deyil",
                examples={
                    "application/json": {
                        "detail": "Invalid credentials"
                    }
                }
            ),
        }
    )
    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data.get('email')
            password = serializer.validated_data.get('password')
            fcm_token = serializer.validated_data.get('fcm_token')
            try:
                user = User.objects.get(email=email, password=password)
            except User.DoesNotExist:
                return Response({"detail": "Invalid credentials"}, status=400)
            if user.status != 'active':
                return Response({"detail": "User is not active"}, status=400)
            if fcm_token:
                user.firebase_token = fcm_token
                user.save()
            request.session['user_id'] = user.id
            return Response({"detail": "Login successful", "user_id": user.id})
        return Response(serializer.errors, status=400)


class UserLogoutAPIView(APIView):
    def post(self, request):
        if 'user_id' in request.session:
            del request.session['user_id']
        return Response({"detail": "Logout successful"})


class ActivateUserAPIView(APIView):
    serializer_class = ActivateUserSerializer

    @swagger_auto_schema(
        operation_summary="User-i aktivləşdirmək",
        operation_description=(
            "Bu endpoint user qeydiyyatdan keçdikdə göndərilən `otp_code` və `email` üzrə user-in "
            "statusunu `active`-ə edir. Səhv `otp_code` gələn halda aktivasiya olunmayacaq."
        ),
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email", "otp_code"],
            properties={
                "email": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    format=openapi.FORMAT_EMAIL,
                    description="User-in email ünvanı",
                    example="example@gmail.com"
                ),
                "otp_code": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="User-in SMS/Email vasitəsilə aldığı OTP kodu",
                    example="123456"
                ),
            },
        ),
        responses={
            200: openapi.Response(
                description="Uğurlu cavab",
                examples={
                    "application/json": {
                        "detail": "User activated successfully"
                    }
                }
            ),
            400: openapi.Response(
                description="OTP və ya digər məlumat səhvdir",
                examples={
                    "application/json": {
                        "detail": "Invalid OTP code"
                    }
                }
            ),
            404: openapi.Response(
                description="Belə user tapılmadı",
                examples={
                    "application/json": {
                        "detail": "User not found"
                    }
                }
            ),
        }
    )
    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data.get('email')
            otp_code = serializer.validated_data.get('otp_code')
            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                return Response({"detail": "User not found"}, status=404)
            if user.otp_code == otp_code:
                user.status = 'active'
                user.otp_code = None
                user.save()
                return Response({"detail": "User activated successfully"})
            else:
                return Response({"detail": "Invalid OTP code"}, status=400)
        return Response(serializer.errors, status=400)


@moderator_role_required('can_view_services')
def services_view(request):
    categories = Category.objects.all().order_by('-created_at')
    context = {
        'categories': categories,
    }
    return render(request, "services.html", context)


@method_decorator(name='post', decorator=swagger_auto_schema(
    operation_summary="Yeni kateqoriya yaratmaq",
    operation_description="Bu endpoint ilə yeni bir kateqoriya yarada bilərsiniz. `name` sahəsi məcburidir, `image` optional-dır.",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["name"],
        properties={
            "name": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Kateqoriyanın adı (məsələn: 'Kosmetologiya')",
                example="Kosmetologiya"
            ),
            "image": openapi.Schema(
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_BINARY,
                description="Kateqoriyaya aid şəkil (optional). Backend file upload qəbul edə bilər.",
                example=None
            ),
        }
    ),
    responses={
        201: openapi.Response(
            description="Kateqoriya uğurla yaradıldı",
            examples={
                "application/json": {
                    "id": 1,
                    "name": "Kosmetologiya",
                    "image": "http://example.com/media/category_image.jpg"
                }
            }
        ),
        400: openapi.Response(
            description="Yanlış məlumat",
            examples={
                "application/json": {
                    "detail": "Bad Request"
                }
            }
        ),
    }
))
class CategoryListCreateAPIView(generics.ListCreateAPIView):
    queryset = Category.objects.all().order_by('-created_at')
    serializer_class = CategorySerializer


class CategoryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class ServiceListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = ServiceSerializer
    queryset = Service.objects.all().order_by('-created_at')

    def get_queryset(self):
        queryset = super().get_queryset()
        category_id = self.request.query_params.get('category', None)
        if category_id is not None:
            queryset = queryset.filter(category_id=category_id)
        return queryset


class ServiceRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer


class ServiceFilterAPIView(generics.ListAPIView):
    serializer_class = ServiceSerializer

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'name',
                openapi.IN_QUERY,
                description="Xidmətin adı (məsələn: 'Lazer epilyasiya'). Oxşar yazı daxil edildikdə belə uyğun nəticələr qaytarılacaq.",
                type=openapi.TYPE_STRING,
                example="Lazer epilyasiya"
            )
        ],
        responses={200: ServiceSerializer(many=True)}
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        name = self.request.query_params.get('name', None)
        queryset = Service.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)
        return queryset


@moderator_role_required('can_view_doctors')
def doctors_view(request):
    doctors = Doctor.objects.all().order_by('-id')
    services = Service.objects.all().order_by('name')

    doctors_list = []
    for doc in doctors:
        schedules = doc.schedules.order_by('day_of_week')
        doc.schedules_by_day = schedules
        doctors_list.append(doc)

    context = {
        'doctors': doctors_list,
        'services': services,
    }
    return render(request, "doctors.html", context)


class DoctorListCreateAPIView(generics.ListCreateAPIView):
    queryset = Doctor.objects.all().order_by('-id')

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DoctorCreateUpdateSerializer
        return DoctorSerializer


class DoctorRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Doctor.objects.all()

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return DoctorCreateUpdateSerializer
        return DoctorSerializer


class DoctorPermissionAPIView(APIView):
    def get_object(self, doctor_id):
        return DoctorPermission.objects.get(doctor_id=doctor_id)

    def patch(self, request, doctor_id):
        try:
            perm = self.get_object(doctor_id)
        except DoctorPermission.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = DoctorPermissionSerializer(
            perm, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DoctorScheduleListCreateAPIView(APIView):
    def get(self, request, doctor_id):
        schedules = DoctorSchedule.objects.filter(
            doctor_id=doctor_id).order_by('day_of_week')
        serializer = DoctorScheduleSerializer(schedules, many=True)
        return Response(serializer.data)


class DoctorScheduleDetailAPIView(APIView):
    def get_object(self, doctor_id, pk):
        return DoctorSchedule.objects.get(doctor_id=doctor_id, pk=pk)

    def patch(self, request, doctor_id, pk):
        schedule = self.get_object(doctor_id, pk)
        serializer = DoctorScheduleSerializer(
            schedule, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, doctor_id, pk):
        schedule = self.get_object(doctor_id, pk)
        serializer = DoctorScheduleSerializer(schedule, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, doctor_id, pk):
        schedule = self.get_object(doctor_id, pk)
        schedule.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DoctorScheduleByDayAPIView(APIView):
    def get_object(self, doctor_id, day_of_week):
        return DoctorSchedule.objects.get(doctor_id=doctor_id, day_of_week=day_of_week)

    def patch(self, request, doctor_id, day_of_week):
        try:
            schedule = self.get_object(doctor_id, day_of_week)
        except DoctorSchedule.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = DoctorScheduleSerializer(
            schedule, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@moderator_role_required('can_view_laser')
def laser_view(request):
    from .models import LaserUsage, Doctor
    import datetime

    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    laser_usages = LaserUsage.objects.all().order_by('-created_at')

    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(
                start_date_str, '%Y-%m-%d').date()
            laser_usages = laser_usages.filter(
                created_at__date__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(
                end_date_str, '%Y-%m-%d').date()
            laser_usages = laser_usages.filter(created_at__date__lte=end_date)
        except ValueError:
            pass

    from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
    page = request.GET.get('page', 1)
    paginator = Paginator(laser_usages, 10)

    try:
        laser_page = paginator.page(page)
    except PageNotAnInteger:
        laser_page = paginator.page(1)
    except EmptyPage:
        laser_page = paginator.page(paginator.num_pages)

    doctors = Doctor.objects.all().order_by('first_name')

    context = {
        'laser_usages': laser_page,
        'doctors': doctors,
        'page_obj': laser_page,
        'paginator': paginator,
        'start_date_str': start_date_str,
        'end_date_str': end_date_str,
    }
    return render(request, "laser.html", context)


class LaserUsageListCreateAPIView(generics.ListCreateAPIView):
    queryset = LaserUsage.objects.all().order_by('-created_at')
    serializer_class = LaserUsageSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)

        moderator_id = request.session.get('moderator_id', None)
        if moderator_id:
            try:
                moderator = Moderator.objects.get(id=moderator_id)
                HistoryLog.objects.create(
                    moderator=moderator,
                    action_type='laser_changed'
                )
            except Moderator.DoesNotExist:
                pass

        return response


class LaserUsageRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LaserUsage.objects.all()
    serializer_class = LaserUsageSerializer


@moderator_role_required('can_view_treatment')
def treatment_view(request):
    services = Service.objects.all().order_by('name')
    treatments = Treatment.objects.all().order_by('-id')
    context = {
        'services': services,
        'treatments': treatments,
    }
    return render(request, 'treatment.html', context)


@method_decorator(name='post', decorator=swagger_auto_schema(
    operation_summary="Yeni müalicə yaratmaq",
    operation_description=(
        "Bu endpoint ilə yeni müalicə yaradıla bilər. `service` ID-sini və `steps` array-ni düzgün daxil etmək lazımdır."
    ),
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["service", "steps"],
        properties={
            "service": openapi.Schema(
                type=openapi.TYPE_INTEGER,
                description="Müalicənin aid olduğu service-in ID-si",
                example=1
            ),
            "steps": openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "title": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            example="Prosesə hazırlıq"
                        ),
                        "description": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            example="İlk addımda pasiyentin dərisinin dezinfeksiyası aparılır"
                        ),
                        "time_offset": openapi.Schema(
                            type=openapi.TYPE_INTEGER,
                            example=20
                        )
                    }
                )
            ),
        }
    ),
    responses={
        201: openapi.Response(
            description="Müalicə uğurla yaradıldı",
            examples={
                "application/json": {
                    "id": 5,
                    "service": 1,
                    "steps": [
                        {
                            "title": "Prosesə hazırlıq",
                            "description": "İlk addımda pasiyentin dərisinin dezinfeksiyası aparılır",
                            "time_offset": 20
                        }
                    ]
                }
            }
        ),
        400: openapi.Response(
            description="Yanlış məlumat",
            examples={
                "application/json": {
                    "detail": "Bad Request"
                }
            }
        ),
    }
))
class TreatmentListCreateAPIView(generics.ListCreateAPIView):
    queryset = Treatment.objects.all().order_by('-id')
    serializer_class = TreatmentSerializer


class TreatmentRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Treatment.objects.all()
    serializer_class = TreatmentSerializer


@moderator_role_required('can_view_discount_code')
def discount_code_view(request):
    search_query = request.GET.get('search', '')
    discount_list = DiscountCode.objects.all().order_by('-created_at')

    if search_query:
        discount_list = discount_list.filter(
            Q(code__icontains=search_query)
        )

    paginator = Paginator(discount_list, 5)
    page = request.GET.get('page')

    try:
        discounts = paginator.page(page)
    except PageNotAnInteger:
        discounts = paginator.page(1)
    except EmptyPage:
        discounts = paginator.page(paginator.num_pages)

    context = {
        'discounts': discounts,
        'search_query': search_query,
    }
    return render(request, "discount-code.html", context)


class DiscountCodeListCreateAPIView(generics.ListCreateAPIView):
    queryset = DiscountCode.objects.all().order_by('-created_at')
    serializer_class = DiscountCodeSerializer
    pagination_class = None


class DiscountCodeRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DiscountCode.objects.all()
    serializer_class = DiscountCodeSerializer


@moderator_role_required('can_view_pages')
def pages_view(request):
    discount_banners = DiscountBanner.objects.all().order_by('-id')
    main_banners = MainBanner.objects.all().order_by('-id')
    services = Service.objects.all().order_by('name')

    context = {
        'discount_banners': discount_banners,
        'main_banners': main_banners,
        'services': services,
    }
    return render(request, 'pages.html', context)


class DiscountBannerListCreateAPIView(generics.ListCreateAPIView):
    queryset = DiscountBanner.objects.all().order_by('-id')
    serializer_class = DiscountBannerSerializer

    def create(self, request, *args, **kwargs):
        existing_count = DiscountBanner.objects.count()
        if existing_count >= 2:
            return Response(
                {"detail": "Maksimum 2 endirim banneri ola bilər."},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().create(request, *args, **kwargs)


class DiscountBannerRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DiscountBanner.objects.all()
    serializer_class = DiscountBannerSerializer


class MainBannerListCreateAPIView(generics.ListCreateAPIView):
    queryset = MainBanner.objects.all().order_by('-id')
    serializer_class = MainBannerSerializer

    def create(self, request, *args, **kwargs):
        existing_count = MainBanner.objects.count()
        if existing_count >= 3:
            return Response(
                {"detail": "Maksimum 3 əsas banner ola bilər."},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().create(request, *args, **kwargs)


class MainBannerRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = MainBanner.objects.all()
    serializer_class = MainBannerSerializer


@moderator_role_required('can_view_calendar')
def reservations_view(request):
    return render(request, "calendar.html")


class ReservationListCreateAPIView(generics.ListCreateAPIView):
    queryset = Reservation.objects.all().order_by('-date', '-start_time')
    serializer_class = ReservationSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        date_param = self.request.query_params.get('date')
        doctor_id = self.request.query_params.get('doctor_id')
        user_id = self.request.query_params.get('user_id')

        if date_param:
            qs = qs.filter(date=date_param)
        if doctor_id:
            qs = qs.filter(doctor_id=doctor_id)
        if user_id:
            qs = qs.filter(user_id=user_id)
        return qs

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)

        moderator_id = request.session.get('moderator_id', None)
        if moderator_id:
            try:
                moderator = Moderator.objects.get(id=moderator_id)
                HistoryLog.objects.create(
                    moderator=moderator,
                    action_type='reservation_created'
                )
            except Moderator.DoesNotExist:
                pass

        return response


class ReservationRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Reservation.objects.all()
    serializer_class = ReservationSerializer


@moderator_role_required('can_view_cashback')
def cashback_view(request):
    current_tab = request.GET.get('currentTab', 'history')

    search_history = request.GET.get('search_history', '')
    history_queryset = CashbackHistory.objects.select_related(
        'user_cashback__user').order_by('-created_at')
    if search_history:
        history_queryset = history_queryset.filter(
            Q(user_cashback__user__first_name__icontains=search_history) |
            Q(user_cashback__user__last_name__icontains=search_history)
        )

    page_history = request.GET.get('page_history', 1)
    paginator_history = Paginator(history_queryset, 5)
    try:
        history_cashbacks = paginator_history.page(page_history)
    except PageNotAnInteger:
        history_cashbacks = paginator_history.page(1)
    except EmptyPage:
        history_cashbacks = paginator_history.page(paginator_history.num_pages)

    search_users = request.GET.get('search_users', '')
    users_queryset = User.objects.all().order_by('-created_at')
    if search_users:
        users_queryset = users_queryset.filter(
            Q(first_name__icontains=search_users) |
            Q(last_name__icontains=search_users)
        )

    for user in users_queryset:
        UserCashback.objects.get_or_create(user=user)

    page_users = request.GET.get('page_users', 1)
    paginator_users = Paginator(users_queryset, 5)
    try:
        user_page = paginator_users.page(page_users)
    except PageNotAnInteger:
        user_page = paginator_users.page(1)
    except EmptyPage:
        user_page = paginator_users.page(paginator_users.num_pages)

    search_budget = request.GET.get('search_budget', '')
    budget_queryset = User.objects.all().order_by('-created_at')
    if search_budget:
        budget_queryset = budget_queryset.filter(
            Q(first_name__icontains=search_budget) |
            Q(last_name__icontains=search_budget)
        )

    for user in budget_queryset:
        UserCashback.objects.get_or_create(user=user)

    page_budget = request.GET.get('page_budget', 1)
    paginator_budget = Paginator(budget_queryset, 5)
    try:
        budget_page = paginator_budget.page(page_budget)
    except PageNotAnInteger:
        budget_page = paginator_budget.page(1)
    except EmptyPage:
        budget_page = paginator_budget.page(paginator_budget.num_pages)

    context = {
        'currentTab': current_tab,

        'search_history': search_history,
        'history_cashbacks': history_cashbacks,

        'search_users': search_users,
        'user_page': user_page,

        'search_budget': search_budget,
        'budget_page': budget_page,
    }
    return render(request, "cashback.html", context)


@swagger_auto_schema(
    method='post',
    operation_summary="User-in cashback statusunu dəyişdirmək (aktiv/deaktiv)",
    operation_description=(
        "Bu endpoint user-in mövcud cashback statusunu dəyişdirir: "
        "əgər `is_active=True` idisə `False` olur."
    ),
    responses={
        200: openapi.Response(
            description="Status uğurla dəyişdirildi",
            examples={
                "application/json": {
                    "status": "ok",
                    "is_active": True
                }
            }
        ),
        404: openapi.Response(
            description="UserCashback tapılmadı",
            examples={
                "application/json": {
                    "error": "UserCashback not found"
                }
            }
        ),
    }
)
@api_view(['POST'])
def toggle_cashback_status(request, user_id):
    try:
        user_cb = UserCashback.objects.get(user_id=user_id)
        user_cb.is_active = not user_cb.is_active
        user_cb.save()
        return Response({'status': 'ok', 'is_active': user_cb.is_active})
    except UserCashback.DoesNotExist:
        return Response({'error': 'UserCashback not found'}, status=404)


@swagger_auto_schema(
    method='post',
    operation_summary="User-in cashback balansını yeniləmək",
    operation_description="Bu endpoint vasitəsilə user-in cashback balansının dəyəri dəyişdirilir.",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["balance"],
        properties={
            "balance": openapi.Schema(
                type=openapi.TYPE_NUMBER,
                description="Yeni cashback dəyəri (məsələn: 120)",
                example=100
            )
        }
    ),
    responses={
        200: openapi.Response(
            description="Balans uğurla yeniləndi",
            examples={
                "application/json": {
                    "status": "ok",
                    "new_balance": 150,
                    "changed_amount": 50
                }
            }
        ),
        404: openapi.Response(
            description="Belə bir istifadəçi üçün Cashback qeydi tapılmadı",
            examples={
                "application/json": {
                    "error": "UserCashback not found"
                }
            }
        ),
        400: openapi.Response(
            description="Düzgün balans verilməyib",
            examples={
                "application/json": {
                    "error": "No balance provided"
                }
            }
        ),
    }
)
@api_view(['POST'])
def update_cashback_balance(request, user_id):
    balance = request.data.get('balance', None)
    if balance is None:
        return Response({'error': 'No balance provided'}, status=400)

    try:
        new_balance = float(balance)
    except ValueError:
        return Response({'error': 'Invalid balance'}, status=400)

    try:
        user_cb = UserCashback.objects.get(user_id=user_id)
    except UserCashback.DoesNotExist:
        return Response({'error': 'UserCashback not found'}, status=404)

    old_balance = float(user_cb.balance)
    diff = new_balance - old_balance

    user_cb.balance = new_balance
    user_cb.save()

    CashbackHistory.objects.create(
        user_cashback=user_cb,
        change_amount=diff
    )

    return Response({
        'status': 'ok',
        'new_balance': user_cb.balance,
        'changed_amount': diff
    })


class UserCashbackListAPIView(generics.ListAPIView):
    queryset = UserCashback.objects.all()
    serializer_class = UserCashbackSerializer


class UserCashbackDetailAPIView(generics.RetrieveAPIView):
    serializer_class = UserCashbackDetailSerializer
    lookup_field = 'user_id'
    lookup_url_kwarg = 'user_id'
    queryset = UserCashback.objects.all()

    def get_object(self):
        user_id = self.kwargs.get(self.lookup_url_kwarg)
        try:
            return UserCashback.objects.get(user_id=user_id)
        except UserCashback.DoesNotExist:
            return None

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj is None:
            return Response([], status=status.HTTP_200_OK)
        serializer = self.get_serializer(obj)
        return Response(serializer.data)


class CashbackHistoryListAPIView(generics.ListAPIView):
    queryset = CashbackHistory.objects.all().order_by('-created_at')
    serializer_class = CashbackHistorySerializer

    def get_object(self):
        user_id = self.kwargs['user_id']
        return get_object_or_404(UserCashback, user__id=user_id)


@moderator_role_required('can_view_payment_acceptance')
def payment_acceptance_view(request):
    from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
    pending_reservations_list = Reservation.objects.filter(
        status='pending').order_by('-created_at')
    accepted_reservations_list = Reservation.objects.filter(
        status='customer_accepted').order_by('-created_at')

    pending_page = request.GET.get('pending_page', 1)
    accepted_page = request.GET.get('accepted_page', 1)
    pending_paginator = Paginator(pending_reservations_list, 10)
    accepted_paginator = Paginator(accepted_reservations_list, 10)
    try:
        pending_reservations = pending_paginator.page(pending_page)
    except (PageNotAnInteger, EmptyPage):
        pending_reservations = pending_paginator.page(1)
    try:
        accepted_reservations = accepted_paginator.page(accepted_page)
    except (PageNotAnInteger, EmptyPage):
        accepted_reservations = accepted_paginator.page(1)

    context = {
        'pending_reservations': pending_reservations,
        'accepted_reservations': accepted_reservations,
    }
    return render(request, "payment-acceptance.html", context)


class AcceptCustomerReservationAPIView(APIView):
    def post(self, request, reservation_id):
        try:
            reservation = Reservation.objects.get(
                id=reservation_id, status='pending')
        except Reservation.DoesNotExist:
            return Response({"detail": "Reservation not found or already accepted."}, status=status.HTTP_404_NOT_FOUND)
        reservation.status = 'customer_accepted'
        reservation.save()
        return Response({"detail": "Reservation accepted."}, status=status.HTTP_200_OK)


class ReservationDetailAPIView(APIView):
    def get(self, request, reservation_id):
        try:
            reservation = Reservation.objects.get(
                id=reservation_id, status='customer_accepted')
        except Reservation.DoesNotExist:
            return Response({"detail": "Reservation not found."}, status=status.HTTP_404_NOT_FOUND)

        if reservation.user:
            customer_name = f"{reservation.user.first_name} {reservation.user.last_name}"
            customer_image = reservation.user.image.url if reservation.user.image else static(
                'images/user-default.jpg')
            try:
                cashback_balance = float(reservation.user.cashback.balance)
            except Exception:
                cashback_balance = 0
        else:
            customer_name = reservation.full_name
            customer_image = static('images/user-default.jpg')
            cashback_balance = 0

        total_price = sum(
            [s.price for s in reservation.services.all()], Decimal('0'))
        from django.db.models import Sum
        payments_sum = reservation.payments.aggregate(
            total=Sum('amount'))['total'] or Decimal('0')
        remaining_debt = max(total_price - payments_sum, Decimal('0'))

        data = {
            "reservation_id": reservation.id,
            "customer_name": customer_name,
            "customer_image": customer_image,
            "cashback_balance": cashback_balance,
            "services": [{"name": service.name, "price": float(service.price)} for service in reservation.services.all()],
            "total_price": float(total_price),
            "remaining_debt": float(remaining_debt),
        }
        return Response(data, status=status.HTTP_200_OK)


class PaymentCreateAPIView(APIView):
    def post(self, request):
        reservation_id = request.data.get("reservation_id")
        payment_type = request.data.get("payment_type")
        amount = request.data.get("amount")
        if not (reservation_id and payment_type and amount):
            return Response({"detail": "Missing fields."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            amount = float(amount)
        except ValueError:
            return Response({"detail": "Invalid amount."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            reservation = Reservation.objects.get(id=reservation_id)
        except Reservation.DoesNotExist:
            return Response({"detail": "Reservation not found."}, status=status.HTTP_404_NOT_FOUND)

        total_price = sum([float(s.price) for s in reservation.services.all()])
        from django.db.models import Sum
        payments_sum = reservation.payments.aggregate(
            total=Sum('amount'))['total'] or 0
        remaining_debt = total_price - float(payments_sum)
        if amount > remaining_debt:
            return Response(
                {"detail": "Ümumi ödəniş qalıq borcdan çox ola bilməz!"},
                status=status.HTTP_400_BAD_REQUEST
            )

        updated_cashback_balance = None
        if payment_type == 'cashback':
            if not reservation.user:
                return Response({"detail": "Cashback ödənişi qeydiyyatsız istifadəçilər üçün mövcud deyil!"}, status=status.HTTP_400_BAD_REQUEST)
            user_cashback = reservation.user.cashback
            if float(user_cashback.balance) < amount:
                return Response({"detail": "Cashback balansı kifayət deyil!"}, status=status.HTTP_400_BAD_REQUEST)
            user_cashback.balance = float(user_cashback.balance) - amount
            user_cashback.save()
            CashbackHistory.objects.create(
                user_cashback=user_cashback, change_amount=-amount)
            updated_cashback_balance = user_cashback.balance

        payment = Payment.objects.create(
            reservation=reservation,
            payment_type=payment_type,
            amount=amount
        )

        payments_sum = reservation.payments.aggregate(
            total=Sum('amount'))['total'] or 0
        if payments_sum >= total_price:
            reservation.status = 'completed'
            reservation.save()

        response_data = {"detail": "Payment recorded.",
                         "payment_id": payment.id}
        if updated_cashback_balance is not None:
            response_data["updated_cashback_balance"] = float(
                updated_cashback_balance)

        if payments_sum >= total_price:
            reservation.status = 'completed'
            reservation.save()

            moderator_id = request.session.get('moderator_id', None)
            if moderator_id:
                try:
                    moderator = Moderator.objects.get(id=moderator_id)
                    HistoryLog.objects.create(
                        moderator=moderator,
                        action_type='reservation_completed'
                    )
                except Moderator.DoesNotExist:
                    pass
        return Response(response_data, status=status.HTTP_201_CREATED)


@receiver(pre_delete, sender=Payment)
def restore_cashback_balance(sender, instance, **kwargs):
    if instance.payment_type == 'cashback':
        reservation = instance.reservation
        if reservation.user and hasattr(reservation.user, 'cashback'):
            user_cashback = reservation.user.cashback
            user_cashback.balance += instance.amount
            user_cashback.save()
            CashbackHistory.objects.create(
                user_cashback=user_cashback, change_amount=instance.amount)


@receiver(post_delete, sender=Payment)
def update_reservation_status_on_payment_delete(sender, instance, **kwargs):
    reservation = instance.reservation
    total_price = sum([float(s.price) for s in reservation.services.all()])
    from django.db.models import Sum
    payments_sum = reservation.payments.aggregate(
        total=Sum('amount'))['total'] or 0
    if payments_sum < total_price and reservation.status == "completed":
        reservation.status = "customer_accepted"
        reservation.save()


@moderator_role_required('can_view_doctor_payment')
def doctor_payment_view(request, doctor_id=None):
    if doctor_id:
        doctor = get_object_or_404(Doctor, id=doctor_id)
    else:
        doctor = Doctor.objects.order_by('-id').first()

    if not doctor:
        return render(request, "doctor-payment.html", {
            'doctor': None,
            'reservations': [],
            'start_date_str': '',
            'end_date_str': '',
            'page_obj': None,
        })

    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    reservations_qs = Reservation.objects.filter(
        doctor=doctor,
        status='completed'
    ).order_by('-date', '-id')

    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(
                start_date_str, '%Y-%m-%d').date()
            reservations_qs = reservations_qs.filter(date__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(
                end_date_str, '%Y-%m-%d').date()
            reservations_qs = reservations_qs.filter(date__lte=end_date)
        except ValueError:
            pass

    page_number = request.GET.get('page', 1)
    paginator = Paginator(reservations_qs, 7)
    try:
        reservations_page = paginator.page(page_number)
    except:
        reservations_page = paginator.page(1)

    context = {
        'doctor': doctor,
        'reservations': reservations_page,
        'start_date_str': start_date_str,
        'end_date_str': end_date_str,
        'page_obj': reservations_page,
    }
    return render(request, "doctor-payment.html", context)


class DoctorServiceViewSet(viewsets.GenericViewSet,
                           mixins.ListModelMixin,
                           mixins.UpdateModelMixin):
    queryset = DoctorService.objects.all()
    serializer_class = DoctorServiceSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        doctor_id = self.request.query_params.get('doctor_id')
        if doctor_id:
            qs = qs.filter(doctor_id=doctor_id)
        return qs

    @action(methods=['post'], detail=False, url_path='bulk_update')
    def bulk_update(self, request):
        data = request.data
        for item in data:
            ds_id = item.get('id')
            commission = item.get('commission_percentage')
            try:
                ds_obj = DoctorService.objects.get(id=ds_id)
                ds_obj.commission_percentage = commission
                ds_obj.save()
            except DoctorService.DoesNotExist:
                continue
        return Response({"detail": "Bulk update successful."}, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
def doctor_constant_cost_view(request, doctor_id):
    try:
        doctor = Doctor.objects.get(id=doctor_id)
    except Doctor.DoesNotExist:
        return Response({"detail": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response({"constant_cost": str(doctor.constant_cost)})
    if request.method == 'POST':
        cost_str = request.data.get('constant_cost')
        try:
            new_cost = float(cost_str)
        except ValueError:
            return Response({"detail": "Invalid cost value."}, status=status.HTTP_400_BAD_REQUEST)
        doctor.constant_cost = new_cost
        doctor.save()
        return Response({"constant_cost": str(doctor.constant_cost)})


@api_view(['GET'])
def doctor_payment_calculate_view(request, doctor_id):
    try:
        doctor = Doctor.objects.get(id=doctor_id)
    except Doctor.DoesNotExist:
        return Response({"detail": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    avans_str = request.GET.get('avans', '0')
    try:
        avans = float(avans_str)
    except ValueError:
        avans = 0.0

    reservations_qs = Reservation.objects.filter(
        doctor=doctor,
        status='completed'
    )
    import datetime
    if start_date_str:
        try:
            sdate = datetime.datetime.strptime(
                start_date_str, '%Y-%m-%d').date()
            reservations_qs = reservations_qs.filter(date__gte=sdate)
        except:
            pass
    if end_date_str:
        try:
            edate = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            reservations_qs = reservations_qs.filter(date__lte=edate)
        except:
            pass

    total_doctor_earning = 0
    for res in reservations_qs:
        for service in res.services.all():
            try:
                ds = DoctorService.objects.get(doctor=doctor, service=service)
                commission = ds.commission_percentage
            except DoctorService.DoesNotExist:
                commission = 0
            total_doctor_earning += float(service.price) * \
                float(commission) / 100.0

    total_doctor_earning = total_doctor_earning - \
        float(doctor.constant_cost) - avans
    if total_doctor_earning < 0:
        total_doctor_earning = 0

    return Response({"total_payment": round(total_doctor_earning, 2)})


@api_view(['POST'])
def update_doctor_total_earning(request, doctor_id):
    try:
        doctor = Doctor.objects.get(id=doctor_id)
    except Doctor.DoesNotExist:
        return Response({"detail": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

    total_value = request.data.get('total_doctor_earning')
    try:
        total_value = float(total_value)
    except (TypeError, ValueError):
        return Response({"detail": "Invalid total earning value."}, status=status.HTTP_400_BAD_REQUEST)

    doctor.total_doctor_earning = total_value
    doctor.save()
    return Response({"total_doctor_earning": str(doctor.total_doctor_earning)}, status=status.HTTP_200_OK)


@api_view(['GET'])
def doctor_search_view(request):
    q = request.GET.get('q', '').strip().lower()
    doctors_qs = Doctor.objects.all().order_by('-id')
    if q:
        doctors_qs = doctors_qs.filter(
            Q(first_name__icontains=q) | Q(
                last_name__icontains=q) | Q(specialty__icontains=q)
        )
    data = []
    for doc in doctors_qs:
        full_name = f"{doc.first_name} {doc.last_name}"
        img_url = doc.image.url if doc.image else '/static/images/user-default.jpg'
        data.append({
            "id": doc.id,
            "name": full_name,
            "image": img_url
        })
    return Response(data, status=200)


@moderator_role_required('can_view_excel')
def excel_view(request):
    search_query = request.GET.get('search', '')
    reservations = Reservation.objects.filter(
        status='completed').order_by('-date', '-start_time')
    if search_query:
        reservations = reservations.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(doctor__first_name__icontains=search_query) |
            Q(doctor__last_name__icontains=search_query)
        )
    paginator = Paginator(reservations, 10)
    page = request.GET.get('page', 1)
    try:
        reservations_page = paginator.page(page)
    except PageNotAnInteger:
        reservations_page = paginator.page(1)
    except EmptyPage:
        reservations_page = paginator.page(paginator.num_pages)
    context = {
        'reservations': reservations_page,
        'search_query': search_query,
        'paginator': paginator,
    }
    return render(request, 'excel.html', context)


def export_excel_view(request):
    search_query = request.GET.get('search', '')
    ids = request.GET.get('ids', None)

    if ids:
        id_list = [int(i) for i in ids.split(',') if i.isdigit()]
        reservations = Reservation.objects.filter(
            id__in=id_list, status='completed').order_by('-date', '-start_time')
    else:
        reservations = Reservation.objects.filter(
            status='completed').order_by('-date', '-start_time')
        if search_query:
            reservations = reservations.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(full_name__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(doctor__first_name__icontains=search_query) |
                Q(doctor__last_name__icontains=search_query)
            )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rezervasiyalar"

    headers = ["Sıra", "Ad", "Nömrə", "Həkim",
               "Qiymət", "Tarix", "Xidmət", "Giriş", "Çıxış"]
    ws.append(headers)

    for index, reservation in enumerate(reservations, start=1):
        if reservation.user:
            ad = f"{reservation.user.first_name} {reservation.user.last_name}"
            nomre = reservation.user.phone
        else:
            ad = reservation.full_name or ""
            nomre = reservation.phone or ""
        hekim = ""
        if reservation.doctor:
            hekim = f"{reservation.doctor.first_name} {reservation.doctor.last_name}"
        qiymet = float(reservation.total_payments)
        tarix = reservation.last_payment_date.strftime(
            '%d/%m/%Y') if reservation.last_payment_date else ""
        xidmet = ", ".join(
            [service.name for service in reservation.services.all()])
        giris = reservation.accepted_at.strftime(
            '%H:%M') if reservation.accepted_at else ""
        cixis = reservation.exit_time.strftime(
            '%H:%M') if reservation.exit_time else ""

        row = [index, ad, nomre, hekim, qiymet, tarix, xidmet, giris, cixis]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reservations_export.xlsx"'
    wb.save(response)
    return response


class AcceptCustomerReservationAPIView(APIView):
    def post(self, request, reservation_id):
        try:
            reservation = Reservation.objects.get(
                id=reservation_id, status='pending')
        except Reservation.DoesNotExist:
            return Response({"detail": "Reservation not found or already accepted."},
                            status=status.HTTP_404_NOT_FOUND)
        reservation.status = 'customer_accepted'
        reservation.accepted_at = timezone.now()
        reservation.save()
        return Response({"detail": "Reservation accepted."}, status=status.HTTP_200_OK)


class ExcelReservationListAPIView(APIView):
    def get(self, request):
        search_query = request.GET.get('search', '')
        reservations = Reservation.objects.filter(
            status='completed').order_by('-date', '-start_time')
        if search_query:
            reservations = reservations.filter(
                Q(user__first_name__icontains=search_query) | Q(user__last_name__icontains=search_query) |
                Q(doctor__first_name__icontains=search_query) | Q(
                    doctor__last_name__icontains=search_query)
            )
        serializer = ExcelReservationSerializer(reservations, many=True)
        return Response(serializer.data)


@moderator_role_required('can_view_statistics')
def statistics_view(request):
    return render(request, "statistics.html")


class StatisticsAPIView(APIView):
    def get(self, request):
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        days_str = request.GET.get('days')

        now = timezone.now().date()
        if start_date_str and end_date_str:
            try:
                start_date = datetime.datetime.strptime(
                    start_date_str, '%Y-%m-%d').date()
                end_date = datetime.datetime.strptime(
                    end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format"}, status=400)
        elif days_str:
            try:
                days = int(days_str)
            except:
                days = 31
            end_date = now
            start_date = now - datetime.timedelta(days=days)
        else:
            end_date = now
            start_date = now - datetime.timedelta(days=31)

        new_users = User.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date).count()
        reservation_count = Reservation.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date).count()
        cashback_total = User.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date,
                                             cashback__isnull=False).aggregate(total=Sum('cashback__balance'))['total'] or 0
        total_sales = Payment.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date).aggregate(total=Sum('amount'))['total'] or 0

        total_users = User.objects.count()
        total_doctors = Doctor.objects.count()

        return Response({
            "new_users": new_users,
            "reservation_count": reservation_count,
            "total_cashback": float(cashback_total),
            "total_sales": float(total_sales),
            "total_users": total_users,
            "total_doctors": total_doctors
        })


@api_view(['GET'])
def doctor_search_view(request):
    q = request.GET.get('q', '').strip().lower()
    doctors_qs = Doctor.objects.all().order_by('-id')
    if q:
        doctors_qs = doctors_qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(specialty__icontains=q)
        ).order_by('-id')
    data = []
    for doc in doctors_qs:
        full_name = f"{doc.first_name} {doc.last_name}"
        img_url = doc.image.url if doc.image else '/static/images/user-default.jpg'
        data.append({
            "id": doc.id,
            "name": full_name,
            "image": img_url
        })
    return Response(data, status=200)


class StatisticsGraphsAPIView(APIView):
    def get(self, request):
        range_type = request.GET.get("range", "ay")
        now = timezone.now()
        payment_qs = Payment.objects.all()
        incomeLabels = []
        incomeData = []
        incomeSum = 0
        card_total = 0
        cash_total = 0
        cashback_total = 0

        if range_type == "gun":
            year = now.year
            month = now.month
            first_day = datetime.date(year, month, 1)
            last_day = (datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
                        ) if month < 12 else datetime.date(year, 12, 31)
            payment_qs = payment_qs.filter(
                created_at__date__gte=first_day, created_at__date__lte=last_day)
            current_day = first_day
            while current_day <= last_day:
                incomeLabels.append(current_day.strftime("%d/%m"))
                day_total = payment_qs.filter(created_at__date=current_day).aggregate(
                    total=Sum("amount"))["total"] or 0
                incomeData.append(float(day_total))
                for pay in payment_qs.filter(created_at__date=current_day):
                    amt = float(pay.amount)
                    if pay.payment_type == "card":
                        card_total += amt
                    elif pay.payment_type == "cash":
                        cash_total += amt
                    else:
                        cashback_total += amt
                current_day += datetime.timedelta(days=1)
            incomeSum = sum(incomeData)
        elif range_type == "hefte":
            current_date = now.date()
            monday = current_date - \
                datetime.timedelta(days=current_date.weekday())
            sunday = monday + datetime.timedelta(days=6)
            payment_qs = payment_qs.filter(
                created_at__date__gte=monday, created_at__date__lte=sunday)
            constWeekDays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            incomeLabels = constWeekDays
            for i in range(7):
                day = monday + datetime.timedelta(days=i)
                day_total = payment_qs.filter(created_at__date=day).aggregate(
                    total=Sum("amount"))["total"] or 0
                incomeData.append(float(day_total))
                for pay in payment_qs.filter(created_at__date=day):
                    amt = float(pay.amount)
                    if pay.payment_type == "card":
                        card_total += amt
                    elif pay.payment_type == "cash":
                        cash_total += amt
                    else:
                        cashback_total += amt
            incomeSum = sum(incomeData)
        else:
            year = now.year
            incomeLabels = ["Jan", "Feb", "Mar", "Apr", "May",
                            "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            for month in range(1, 13):
                first_day = datetime.date(year, month, 1)
                if month == 12:
                    last_day = datetime.date(year, 12, 31)
                else:
                    last_day = datetime.date(
                        year, month + 1, 1) - datetime.timedelta(days=1)
                month_total = payment_qs.filter(created_at__date__gte=first_day, created_at__date__lte=last_day).aggregate(
                    total=Sum("amount"))["total"] or 0
                incomeData.append(float(month_total))
                for pay in payment_qs.filter(created_at__date__gte=first_day, created_at__date__lte=last_day):
                    amt = float(pay.amount)
                    if pay.payment_type == "card":
                        card_total += amt
                    elif pay.payment_type == "cash":
                        cash_total += amt
                    else:
                        cashback_total += amt
            incomeSum = sum(incomeData)

        return Response({
            "incomeLabels": incomeLabels,
            "incomeData": incomeData,
            "incomeSum": round(incomeSum, 2),
            "paymentMethods": {
                "card": round(card_total, 2),
                "cash": round(cash_total, 2),
                "cashback": round(cashback_total, 2)
            }
        })


class DoctorEarningsAPIView(APIView):
    def get(self, request):
        doctor_id = request.GET.get("doctor_id")
        if not doctor_id:
            return Response({"error": "doctor_id required"}, status=400)
        try:
            doctor = Doctor.objects.get(id=doctor_id)
        except Doctor.DoesNotExist:
            return Response({"error": "Doctor not found"}, status=404)
        year = timezone.now().year
        labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        earnings = []
        for month in range(1, 13):
            first_day = datetime.date(year, month, 1)
            if month == 12:
                last_day = datetime.date(year, 12, 31)
            else:
                last_day = datetime.date(
                    year, month + 1, 1) - datetime.timedelta(days=1)
            reservations = Reservation.objects.filter(
                doctor=doctor,
                status="completed",
                date__gte=first_day,
                date__lte=last_day
            )
            month_total = Payment.objects.filter(
                reservation__in=reservations,
                created_at__date__gte=first_day,
                created_at__date__lte=last_day
            ).aggregate(total=Sum("amount"))["total"] or 0
            earnings.append(float(month_total))
        return Response({
            "labels": labels,
            "earnings": earnings
        })


@moderator_role_required('can_view_moderator')
def moderator_view(request):
    search_query = request.GET.get('search', '')
    moderator_list = Moderator.objects.all().order_by('-created_at')

    if search_query:
        search_lower = search_query.lower()
        if search_lower in ['aktiv', 'active']:
            moderator_list = moderator_list.filter(status='active')
        elif search_lower in ['deaktiv', 'inactive']:
            moderator_list = moderator_list.filter(status='inactive')
        else:
            moderator_list = moderator_list.filter(
                full_name__icontains=search_query)

    page = request.GET.get('page', 1)
    paginator = Paginator(moderator_list, 10)
    try:
        moderators = paginator.page(page)
    except PageNotAnInteger:
        moderators = paginator.page(1)
    except EmptyPage:
        moderators = paginator.page(paginator.num_pages)

    context = {
        'moderators': moderators,
        'search_query': search_query,
    }
    return render(request, "moderator.html", context)


class ModeratorListCreateAPIView(generics.ListCreateAPIView):
    queryset = Moderator.objects.all().order_by('-created_at')
    serializer_class = ModeratorSerializer


class ModeratorRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Moderator.objects.all()
    serializer_class = ModeratorSerializer


@moderator_role_required('can_view_history')
def history_view(request):
    history_logs = HistoryLog.objects.select_related(
        'moderator').order_by('-created_at')

    page = request.GET.get('page', 1)
    paginator = Paginator(history_logs, 10)

    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)

    context = {
        'logs': logs_page,
    }
    return render(request, "history.html", context)


@moderator_role_required('can_view_notification')
def notification_view(request):
    search_query = request.GET.get('search', '')
    gender_filter = request.GET.get('gender', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    user_list = User.objects.all().order_by('-created_at')

    if search_query:
        user_list = user_list.filter(
            Q(first_name__icontains=search_query) | Q(
                last_name__icontains=search_query)
        )
    if gender_filter in ['male', 'female']:
        user_list = user_list.filter(gender=gender_filter)

    import datetime
    if start_date_str:
        try:
            start_date = datetime.datetime.strptime(
                start_date_str, '%Y-%m-%d').date()
            user_list = user_list.filter(created_at__date__gte=start_date)
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.datetime.strptime(
                end_date_str, '%Y-%m-%d').date()
            user_list = user_list.filter(created_at__date__lte=end_date)
        except ValueError:
            pass

    page = request.GET.get('page', 1)
    paginator = Paginator(user_list, 10)
    try:
        users_page = paginator.page(page)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)

    context = {
        'users': users_page,
        'search_query': search_query,
    }
    return render(request, "notification.html", context)


class NotificationListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = NotificationSerializer

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'id',
                openapi.IN_QUERY,
                description="Bildirişlərin aid olduğu istifadəçi id-ləri, vergüllə ayrılmış. Məsələn: 1,2,3",
                type=openapi.TYPE_STRING,
                required=False
            )
        ],
        operation_summary="Bildirişlərin siyahısı",
        operation_description=(
            "GET metodu: Əgər query parameter olaraq 'id' verilsə, həmin id-lərə sahib istifadəçilərin "
            "bildirişlərini qaytarır, verilmədikdə isə bütün bildirişləri qaytarır.\n\n"
            "POST metodu: 'user_ids' və 'message' sahələri vasitəsilə yeni bildiriş yaradır."
        ),
        responses={
            200: NotificationSerializer(many=True),
            201: NotificationSerializer(),
            400: "Bad Request"
        }
    )
    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["user_ids", "message"],
            properties={
                "user_ids": openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Schema(type=openapi.TYPE_INTEGER),
                    description="Bildirişin göndəriləcəyi istifadəçi id-ləri"
                ),
                "message": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Bildirişin mətni"
                )
            }
        ),
        responses={
            201: NotificationSerializer(),
            400: "Bad Request"
        },
        operation_summary="Yeni bildiriş yarat",
        operation_description="Verilmiş istifadəçi id-ləri və mesaj əsasında yeni bildiriş yaradır."
    )
    def post(self, request, *args, **kwargs):
        user_ids = request.data.get('user_ids', [])
        message = request.data.get('message', '').strip()
        if not user_ids or not message:
            return Response(
                {"detail": "user_ids və message sahələri zəruridir."},
                status=status.HTTP_400_BAD_REQUEST
            )
        notification_obj = Notification.objects.create(message=message)
        notification_obj.recipients.set(User.objects.filter(id__in=user_ids))

        recipients = notification_obj.recipients.all()
        registration_tokens = list(
            recipients.values_list('firebase_token', flat=True))
        registration_tokens = [token for token in registration_tokens if token]

        if registration_tokens:
            push_response = send_push_notification(
                registration_tokens,
                title="Yeni bildiriş",
                body=message,
                data_message={"notification_id": str(notification_obj.id)}
            )

        serializer = self.get_serializer(notification_obj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def get_queryset(self):
        queryset = Notification.objects.all().order_by('-id')
        recipient_ids = self.request.query_params.get('id', None)
        if recipient_ids:
            id_list = [int(x) for x in recipient_ids.split(',')
                       if x.strip().isdigit()]
            queryset = queryset.filter(recipients__id__in=id_list).distinct()
        return queryset


@api_view(['POST'])
def update_firebase_token(request):
    user_id = request.data.get("user_id")
    new_token = request.data.get("firebase_token")
    if not (user_id and new_token):
        return Response({"detail": "User id və token tələb olunur."}, status=400)

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "Belə user yoxdur."}, status=404)

    user.firebase_token = new_token
    user.save()
    return Response({"detail": "Firebase token yeniləndi."}, status=200)


class ForgotPasswordAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Şifrəni unutdum ilə OTP kod göndərilməsi",
        operation_description="User emailini daxil edir və email mövcuddursa, sistem emailə 6 rəqəmli OTP kod göndərir.",
        request_body=ForgotPasswordSerializer,
        responses={
            200: openapi.Response(
                description="OTP kod emailə göndərildi",
                examples={
                    "application/json": {
                        "detail": "OTP kod email-ə göndərildi."
                    }
                }
            ),
            404: openapi.Response(
                description="Email tapılmadı",
                examples={
                    "application/json": {
                        "detail": "Bu emailə uyğun istifadəçi tapılmadı."
                    }
                }
            ),
        }
    )
    def post(self, request):
        serializer = ForgotPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response(
                {"detail": "Bu emailə uyğun istifadəçi tapılmadı."},
                status=status.HTTP_404_NOT_FOUND
            )

        otp_code = str(random.randint(100000, 999999))
        user.otp_code = otp_code
        user.save()

        subject = "Şifrəni Yenilə - OTP Kodu"
        message = f"Sizin OTP kodunuz: {otp_code}\nBu kodu heç kimlə paylaşmayın."
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [email]
        send_mail(subject, message, from_email, recipient_list)

        return Response({"detail": "OTP kod email-ə göndərildi."}, status=status.HTTP_200_OK)


class VerifyOtpAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="OTP Təsdiqi",
        operation_description="Bu endpoint, istifadəçinin email və OTP kodunu yoxlayır. Əgər OTP kodu düzgün olarsa, istifadəçiyə 'OTP təsdiq edildi' mesajı qaytarılır və session-da email qeyd olunur.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email", "otp_code"],
            properties={
                "email": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    format="email",
                    description="İstifadəçinin email ünvanı",
                    example="example@gmail.com"
                ),
                "otp_code": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="İstifadəçiyə göndərilmiş 6 rəqəmli OTP kodu",
                    example="123456"
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="OTP təsdiqi uğurlu oldu.",
                examples={
                    "application/json": {
                        "detail": "OTP təsdiq edildi. Şifrənizi yeniləyə bilərsiniz."
                    }
                }
            ),
            400: openapi.Response(
                description="OTP kod səhvdir.",
                examples={
                    "application/json": {"detail": "OTP kod yanlışdır."}
                }
            ),
            404: openapi.Response(
                description="Emailə uyğun istifadəçi tapılmadı.",
                examples={
                    "application/json": {"detail": "Bu emailə uyğun istifadəçi tapılmadı."}
                }
            ),
        }
    )
    def post(self, request):
        serializer = VerifyOtpSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']
        otp_code = serializer.validated_data['otp_code']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({"detail": "Bu emailə uyğun istifadəçi tapılmadı."}, status=404)

        if user.otp_code != otp_code:
            return Response({"detail": "OTP kod yanlışdır."}, status=400)

        request.session['otp_verified_email'] = email

        return Response({"detail": "OTP təsdiq edildi. Şifrənizi yeniləyə bilərsiniz."}, status=200)


class UpdatePasswordAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Yeni Şifrə Təyin Et",
        operation_description="Bu endpoint, OTP təsdiqi edilmiş email və yeni şifrəni qəbul edir. Əgər session-da təsdiqlənmiş email varsa, həmin istifadəçinin şifrəsi yenilənir.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email", "new_password"],
            properties={
                "email": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    format="email",
                    description="İstifadəçinin email ünvanı (OTP təsdiqi edilmiş email olmalıdır)",
                    example="example@gmail.com"
                ),
                "new_password": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Yeni şifrə",
                    example="newpassword123"
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="Şifrə uğurla yeniləndi.",
                examples={
                    "application/json": {"detail": "Şifrə uğurla yeniləndi."}
                }
            ),
            400: openapi.Response(
                description="OTP təsdiqi edilməyib və ya səhv məlumat.",
                examples={
                    "application/json": {"detail": "OTP təsdiqi edilməyib."}
                }
            ),
            404: openapi.Response(
                description="Emailə uyğun istifadəçi tapılmadı.",
                examples={
                    "application/json": {"detail": "Bu emailə uyğun istifadəçi tapılmadı."}
                }
            ),
        }
    )
    def post(self, request):
        serializer = UpdatePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']
        new_password = serializer.validated_data['new_password']

        verified_email = request.session.get('otp_verified_email')
        if not verified_email or verified_email != email:
            return Response({"detail": "OTP təsdiqi edilməyib."}, status=400)

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({"detail": "Bu emailə uyğun istifadəçi tapılmadı."}, status=404)

        user.password = new_password
        user.otp_code = None
        user.save()

        del request.session['otp_verified_email']

        return Response({"detail": "Şifrə uğurla yeniləndi."}, status=200)


class AvailableDoctorsAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Seçilmiş tarix və saat intervalında uyğun həkimləri göstərir",
        operation_description=(
            "Bu endpoint, sorğuda göndərilən xidmətlər, tarix və saat intervalına əsaslanaraq, "
            "həmin şərtlərə uyğun və aktiv olan həkimləri filtrləyir. Parametrlər ayrıca da verilə bilər.\n\n"
            "Misal:\n"
            "  - Sadəcə services: `/api/available-doctors/?services=1,2,3`\n"
            "  - Sadəcə date: `/api/available-doctors/?date=2025-02-19`\n"
            "  - Services və date: `/api/available-doctors/?services=1,2&date=2025-02-19`\n"
            "  - Services, date, start_time və end_time: "
            "`/api/available-doctors/?services=1,2,3&date=2025-02-19&start_time=09:00&end_time=10:00`"
        ),
        manual_parameters=[
            openapi.Parameter(
                'services',
                openapi.IN_QUERY,
                description="Xidmətlərin ID-ləri (vergüllə ayrılmış). Məsələn: `services=1,2,3`",
                type=openapi.TYPE_STRING,
                required=False
            ),
            openapi.Parameter(
                'date',
                openapi.IN_QUERY,
                description="Rezervasiya tarixi (YYYY-MM-DD formatında). Məsələn: `2025-02-19`",
                type=openapi.TYPE_STRING,
                required=False
            ),
            openapi.Parameter(
                'start_time',
                openapi.IN_QUERY,
                description="Başlama saatı (HH:MM formatında). Məsələn: `09:00`",
                type=openapi.TYPE_STRING,
                required=False
            ),
            openapi.Parameter(
                'end_time',
                openapi.IN_QUERY,
                description="Bitiş saatı (HH:MM formatında). Məsələn: `10:00`. Verilməzsə, başlama saatından 30 dəqiqə sonra qəbul edilir.",
                type=openapi.TYPE_STRING,
                required=False
            ),
        ],
        responses={
            200: openapi.Response(
                description="Filtrlənmiş həkimlərin siyahısı",
                examples={
                    "application/json": [
                        {
                            "id": 1,
                            "first_name": "Elvin",
                            "last_name": "Hümbətov",
                            "image": "",
                            "specialty": "Dermatoloq",
                            "rating": "4.5",
                            "experience": "5 il",
                            "about": "",
                            "services": ["Lazer", "Krio"],
                            "service_ids": [1, 2],
                            "schedules": [],
                            "permission": {}
                        }
                    ]
                }
            )
        }
    )
    def get(self, request):
        services_param = request.query_params.get('services', '')
        date_str = request.query_params.get('date', '')
        start_time_str = request.query_params.get('start_time', '')
        end_time_str = request.query_params.get('end_time', '')

        doctors_qs = Doctor.objects.all()

        if services_param:
            try:
                service_ids = [int(sid.strip()) for sid in services_param.split(
                    ',') if sid.strip().isdigit()]
                for sid in service_ids:
                    doctors_qs = doctors_qs.filter(services__id=sid)
                doctors_qs = doctors_qs.distinct()
            except Exception:
                pass

        if date_str:
            try:
                date_obj = datetime.datetime.strptime(
                    date_str, '%Y-%m-%d').date()
                weekday = date_obj.weekday()
                day_of_week_val = weekday + 1
                doctors_qs = doctors_qs.filter(
                    schedules__day_of_week=day_of_week_val, schedules__is_active=True).distinct()
            except ValueError:
                date_obj = None
        else:
            date_obj = None

        if start_time_str and date_obj:
            try:
                start_time_obj = datetime.datetime.strptime(
                    start_time_str, '%H:%M').time()
            except ValueError:
                start_time_obj = None
            if start_time_obj:
                if end_time_str:
                    try:
                        end_time_obj = datetime.datetime.strptime(
                            end_time_str, '%H:%M').time()
                    except ValueError:
                        end_time_obj = None
                else:
                    end_time_obj = (datetime.datetime.combine(
                        date_obj, start_time_obj) + datetime.timedelta(minutes=30)).time()
                if end_time_obj and datetime.datetime.combine(date_obj, end_time_obj) <= datetime.datetime.combine(date_obj, start_time_obj):
                    return Response(
                        {"detail": "Bitiş saatı başlama saatından əvvəl ola bilməz."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                start_time_obj = None
                end_time_obj = None
        else:
            start_time_obj = None
            end_time_obj = None

        eligible_doctors = []
        for doc in doctors_qs:
            if date_obj:
                schedule = doc.schedules.filter(
                    day_of_week=day_of_week_val, is_active=True).first()
                if not schedule:
                    continue
                if start_time_obj and end_time_obj:
                    if schedule.start_time > start_time_obj or schedule.end_time < end_time_obj:
                        continue

                perm = getattr(doc, 'permission', None)
                permission_blocks = False
                if perm and perm.permission_type != 'none':
                    if perm.permission_type == 'date_range':
                        if perm.start_date and perm.end_date and perm.start_date <= date_obj <= perm.end_date:
                            permission_blocks = True
                    elif perm.permission_type == 'time_range':
                        if perm.specific_date and perm.specific_date == date_obj:
                            if perm.start_time and perm.end_time and start_time_obj and end_time_obj:
                                req_start_dt = datetime.datetime.combine(
                                    date_obj, start_time_obj)
                                req_end_dt = datetime.datetime.combine(
                                    date_obj, end_time_obj)
                                perm_start_dt = datetime.datetime.combine(
                                    date_obj, perm.start_time)
                                perm_end_dt = datetime.datetime.combine(
                                    date_obj, perm.end_time)
                                if req_start_dt < perm_end_dt and req_end_dt > perm_start_dt:
                                    permission_blocks = True
                            else:
                                permission_blocks = True
                if permission_blocks:
                    continue

                if start_time_obj and end_time_obj:
                    existing_reservations = Reservation.objects.filter(
                        doctor=doc, date=date_obj)
                    conflict_found = False
                    req_start_dt = datetime.datetime.combine(
                        date_obj, start_time_obj)
                    req_end_dt = datetime.datetime.combine(
                        date_obj, end_time_obj)
                    for res in existing_reservations:
                        res_start = res.start_time
                        res_end = res.end_time if res.end_time else (datetime.datetime.combine(
                            date_obj, res.start_time) + datetime.timedelta(minutes=30)).time()
                        res_start_dt = datetime.datetime.combine(
                            date_obj, res_start)
                        res_end_dt = datetime.datetime.combine(
                            date_obj, res_end)
                        if req_start_dt < res_end_dt and req_end_dt > res_start_dt:
                            conflict_found = True
                            break
                    if conflict_found:
                        continue

            eligible_doctors.append(doc)

        serializer = DoctorSerializer(
            eligible_doctors,
            many=True,
            context={'request': request}
        )

        return Response(serializer.data, status=status.HTTP_200_OK)


class UserTreatmentsAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="İstifadəçinin müalicə məlumatlarını əldə et",
        operation_description=(
            "Bu endpoint sorğu vasitəsilə göndərilən user id əsasında, "
            "həmin istifadəçinin bütün rezervasiyalarındakı "
            "seçilmiş xidmətlərə aid müalicə məlumatlarını səhifələnmiş şəkildə qaytarır.\n\n"
            "Hər səhifədə yalnız bir rezervasiyaya aid bütün xidmətlərin müalicələri göstərilir və "
            "ən son yaradılan rezervasiya ilk səhifədə yerləşir."
        ),
        manual_parameters=[
            openapi.Parameter(
                'user_id',
                openapi.IN_QUERY,
                description="İstifadəçi ID",
                type=openapi.TYPE_INTEGER,
                required=True,
                example=10
            )
        ],
        responses={
            200: openapi.Response(
                description="Rezervasiyaların səhifələnmiş siyahısı və onların müalicə məlumatları.",
                examples={
                    "application/json": {
                        "count": 3,
                        "next": "http://example.com/api/user-treatments/?page=2",
                        "previous": None,
                        "results": [
                            {
                                "reservation_id": 12,
                                "created_at": "2025-03-01T12:34:56Z",
                                "treatments": [
                                    {
                                        "id": 1,
                                        "service": 3,
                                        "created_at": "2025-03-01T12:00:00Z",
                                        "steps": [
                                            {
                                                "id": 1,
                                                "title": "Prosesə hazırlıq",
                                                "description": "Pasiyentin dərisinin dezinfeksiyası",
                                                "time_offset": 10,
                                                "created_at": "2025-03-01T12:00:00Z"
                                            }
                                        ]
                                    }
                                ]
                            }
                        ]
                    }
                }
            ),
            400: openapi.Response(
                description="User id göndərilməyib və ya səhv məlumat.",
                examples={"application/json": {"detail": "User id required."}}
            ),
            404: openapi.Response(
                description="Bu istifadəçiyə aid rezervasiya tapılmadı.",
                examples={
                    "application/json": {"detail": "No reservations found for this user."}}
            ),
        }
    )
    def get(self, request):
        user_id = request.query_params.get("user_id")
        if not user_id:
            return Response({"detail": "User id required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user_id = int(user_id)
        except ValueError:
            return Response({"detail": "User id must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

        reservations = Reservation.objects.filter(
            user_id=user_id).order_by("-created_at")
        if not reservations.exists():
            reservations = Reservation.objects.none()

        paginator = PageNumberPagination()
        paginator.page_size = 1
        paginated_reservations = paginator.paginate_queryset(
            reservations, request)
        serializer = ReservationTreatmentSerializer(
            paginated_reservations, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)


class ResetPasswordAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Şifrəni dəyişdir",
        operation_description="Bu endpoint istifadəçinin email, cari şifrə və yeni şifrəsini qəbul edir. "
                              "Əgər cari şifrə doğru göstərilərsə, həmin istifadəçinin şifrəsi yeni şifrə ilə dəyişdirilir.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email", "current_password", "new_password"],
            properties={
                "email": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    format="email",
                    description="İstifadəçinin email ünvanı",
                    example="example@gmail.com"
                ),
                "current_password": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="İstifadəçinin cari şifrəsi",
                    example="oldpassword123"
                ),
                "new_password": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="İstifadəçinin yeni şifrəsi (ən az 6 simvol)",
                    example="newpassword123"
                ),
            }
        ),
        responses={
            200: openapi.Response(
                description="Şifrə uğurla dəyişdirildi.",
                examples={
                    "application/json": {"detail": "Şifrə uğurla dəyişdirildi."}}
            ),
            400: openapi.Response(
                description="Cari şifrə yanlışdır və ya digər məlumat səhvdir.",
                examples={
                    "application/json": {"detail": "Cari şifrə yanlışdır."}}
            ),
            404: openapi.Response(
                description="Bu emailə uyğun istifadəçi tapılmadı.",
                examples={
                    "application/json": {"detail": "Bu emailə uyğun istifadəçi tapılmadı."}}
            ),
        }
    )
    def post(self, request):
        serializer = ResetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']
        current_password = serializer.validated_data['current_password']
        new_password = serializer.validated_data['new_password']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({"detail": "Bu emailə uyğun istifadəçi tapılmadı."}, status=404)

        if user.password != current_password:
            return Response({"detail": "Cari şifrə yanlışdır."}, status=400)

        user.password = new_password
        user.save()
        return Response({"detail": "Şifrə uğurla dəyişdirildi."}, status=200)


class CalculateDiscountPercentageAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Discount dəyərinin hesablanması",
        operation_description=(
            "Bu endpoint discount code və bir ədəd (value) qəbul edir. \n\n"
            "Verilən value üzərində discount code-un discount_percentage faizini tətbiq edərək, "
            "endirimli dəyəri hesablayır.\n\n"
            "Məsələn, əgər discount code-un discount_percentage 10%‑dirsə və value 5‑dirsə, "
            "hesablama: 5 - (5 * 10/100) = 4.50."
        ),
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["code", "value"],
            properties={
                "code": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Discount code (məsələn: TEST10)",
                    example="TEST10"
                ),
                "value": openapi.Schema(
                    type=openapi.TYPE_NUMBER,
                    description="Əsas dəyər (məsələn: 5)",
                    example=5
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="Yeni dəyər uğurla hesablanıb.",
                examples={
                    "application/json": {
                        "detail": "Discount applied successfully.",
                        "new_value": "4.50"
                    }
                }
            ),
            400: openapi.Response(
                description="Yanlış məlumat və ya aktiv olmayan discount code.",
                examples={
                    "application/json": {
                        "detail": "Both code and value are required."
                    }
                }
            ),
            404: openapi.Response(
                description="Discount code tapılmadı.",
                examples={
                    "application/json": {
                        "detail": "Discount code not found."
                    }
                }
            ),
        }
    )
    def post(self, request):
        code = request.data.get("code")
        value = request.data.get("value")

        if not code or value is None:
            return Response({"detail": "Both code and value are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            provided_value = Decimal(str(value))
        except Exception:
            return Response({"detail": "Value must be a number."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            discount = DiscountCode.objects.get(code=code)
        except DiscountCode.DoesNotExist:
            return Response({"detail": "Discount code not found."}, status=status.HTTP_404_NOT_FOUND)

        if not discount.is_active:
            return Response({"detail": "Discount code is not active."}, status=status.HTTP_400_BAD_REQUEST)

        discount_fraction = discount.discount_percent / Decimal('100')
        new_value = provided_value - (provided_value * discount_fraction)

        if new_value < 0:
            new_value = Decimal('0')

        return Response({
            "detail": "Discount applied successfully.",
            "new_value": str(new_value.quantize(Decimal('0.01')))
        }, status=status.HTTP_200_OK)


class UserReservationsAPIView(APIView):
    def get(self, request, user_id):
        reservations = Reservation.objects.filter(user_id=user_id).order_by('-date', '-start_time')
        
        paginator = PageNumberPagination()
        paginator.page_size = 5 
        paginated_reservations = paginator.paginate_queryset(reservations, request)
        
        serializer = ReservationDetailSerializer(paginated_reservations, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)
    

class ChangePasswordAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Şifrəni dəyişdir",
        operation_description="Bu endpoint, user_id, old_password və new_password məlumatlarını qəbul edir. Əgər old_password düzgündürsə, şifrə new_password ilə yenilənir.",
        request_body=ChangePasswordSerializer,
        responses={
            200: openapi.Response(
                description="Şifrə uğurla dəyişdirildi.",
                examples={"application/json": {"detail": "Password changed successfully."}}
            ),
            400: openapi.Response(
                description="Köhnə şifrə yanlışdır.",
                examples={"application/json": {"detail": "Old password is incorrect."}}
            ),
            404: openapi.Response(
                description="İstifadəçi tapılmadı.",
                examples={"application/json": {"detail": "User not found."}}
            ),
        }
    )
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user_id = serializer.validated_data['user_id']
        old_password = serializer.validated_data['old_password']
        new_password = serializer.validated_data['new_password']
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        
        if user.password != old_password:
            return Response({"detail": "Old password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)
        
        user.password = new_password
        user.save()
        
        return Response({"detail": "Password changed successfully."}, status=status.HTTP_200_OK)